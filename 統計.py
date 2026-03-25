import ctypes
ctypes.windll.shcore.SetProcessDpiAwareness(1)  # Enable DPI awareness
from ctypes import windll, Structure, c_long, byref
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook # type: ignore
import openpyxl # type: ignore
import os
import shutil
from openpyxl.styles import Alignment, Border, Side # type: ignore

class RECT(Structure):
    _fields_ = [("left", c_long),
                ("top", c_long),
                ("right", c_long),
                ("bottom", c_long)]

def get_work_area_height():
    rect = RECT()
    SPI_GETWORKAREA = 0x0030
    windll.user32.SystemParametersInfoW(SPI_GETWORKAREA, 0, byref(rect), 0)
    return rect.bottom - rect.top

# 宣告視窗且設定全屏
_mainWindow = tk.Tk()
_mainWindow.title("學位服歸還系統")

# Get scaled screen dimensions
scaling_factor = max(1.0, _mainWindow.winfo_fpixels('1i') / 80.0) * 0.8  # Adjusted scaling factor
_screenWidth = _mainWindow.winfo_screenwidth()
_screenHeight = get_work_area_height()
_mainWindow.geometry(f"{_screenWidth}x{_screenHeight}")
_mainWindow.state('zoomed')

# 基本配色
_backgroundColor = "#343541"
_buttonColor = "#1E90FF"

# 全畫面使用 grid，三個 row，全部佔滿
_mainWindow.rowconfigure(0, weight=0, minsize=int(_screenHeight * 0.23))  # top 25%
_mainWindow.rowconfigure(1, weight=0, minsize=int(_screenHeight * 0.64))  # middle 55%
_mainWindow.rowconfigure(2, weight=0, minsize=int(_screenHeight * 0.13)) 
_mainWindow.columnconfigure(0, weight=1)

# 字體設定
base_font_size = 18
scaled_font_size = int(base_font_size / scaling_factor)
font_style = ("微軟正黑體", scaled_font_size)
header_font = ("微軟正黑體", int(scaled_font_size * 1.2), "bold")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center')

# 建立字體樣式
style = ttk.Style()
style.configure("Treeview", font=font_style, rowheight=int(50 / scaling_factor))
style.configure("Treeview.Heading", font=header_font, rowheight=int(50 / scaling_factor))

# 學號綁定變數
ID = tk.StringVar()
_dayVar = tk.StringVar()
_monthVar = tk.StringVar()
_statusOption1 = tk.IntVar()
_statusOption1.set(0)
_statusOptions = ["上午", "下午", "全天"]
_selectedOption = tk.StringVar()
_selectedOption.set(_statusOptions[0])
_collegeNums = [[0 for _ in range(3)] for _ in range(9)]
_totalNums = [0, 0, 0, 0]
_returnStatus = ("未歸還", "已歸還", "遺失", "已歸還(新衣)", "已歸還(報廢)")
_deleteStatus = ("歸還", "遺失", "新衣", "報廢")
_collegeName = ("文學院", "理學院", "社科院", "工學院", "管學院", "法學院", "教育學院", None, "紫荊不分系")

def create_path(pathname):
    if not os.path.isdir(pathname):
        os.mkdir(pathname)

def resetNums():
    for i in range(3):
        _totalNums[i] = 0
        for j in range(9):
            _collegeNums[j][i] = 0

def init_excel():
    global _allNameList, _targetSheet
    if not os.path.exists("所有系所名單.xlsx"):
        show_customWindow("找不到所有系所名單", "notice")
        _allNameList = None
        _targetSheet = None
        return
    _allNameList = load_workbook("所有系所名單.xlsx")
    _targetSheet = _allNameList.active

def enter(self):
    search_studentID()

def search_studentID(_targetRecord):
    global _targetRow
    _targetRow = -1
    _status = "未借用"
    found = False
    for row in range(1, _targetSheet.max_row+1):
        _studentID_value = str(_targetSheet.cell(row=row, column=1).value).strip()
        if ID.get() == _studentID_value:
            _targetRow = row
            _name = _targetSheet.cell(row=_targetRow, column=2).value
            _department = get_deparmentInfo(_targetSheet.cell(row=_targetRow, column=3).value, "name")
            _time = _targetSheet.cell(row=_targetRow, column=7).value
            if _time is None:
                _time = ''
            else:
                if isinstance(_time, datetime):
                    _time = _time.strftime('%Y/%m/%d %H:%M:%S')
            _tempStatus = _targetSheet.cell(row=_targetRow, column=6).value
            if _targetSheet.cell(row=_targetRow, column=5).value == 1: 
                if _tempStatus == 0 or _tempStatus is None:
                    _status = "借用 / " + _returnStatus[0]
                elif _tempStatus == 1:
                    _status = "借用 / " + _returnStatus[1]
                elif _tempStatus == 2:
                    _status = "借用 / " + _returnStatus[2]
                elif _tempStatus == 3:
                    _status = "借用 / " + _returnStatus[3]
                elif _tempStatus == 4:
                    _status = "借用 / " + _returnStatus[4]
            for row in _targetRecord.get_children():
                _targetRecord.delete(row)
            _targetRecord.insert('', 'end', values=(_studentID_value, _name, _department, _status, _time))
            found = True
            break
    if not found:
        show_customWindow("查無此人！", "error")
        return

def change_return_status(_targetRecord, _historyRecords, _status):
    if _targetRow == -1:
        show_customWindow("請輸入學號！", "error")
        return
    _borrowCell = _targetSheet.cell(row=_targetRow, column=5).value
    _studentID_value = str(_targetSheet.cell(row=_targetRow, column=1).value).strip()
    _name = _targetSheet.cell(row=_targetRow, column=2).value
    _department = get_deparmentInfo(_targetSheet.cell(row=_targetRow, column=3).value, "name")
    _time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if _borrowCell is None:
        if _status == 0:
            for row in _targetRecord.get_children():
                _targetRecord.delete(row)
            _targetSheet.cell(row=_targetRow, column=5).value = 1
            _targetSheet.cell(row=_targetRow, column=6).value = _status
            _targetRecord.insert('', 'end', values=(_studentID_value, _name, _department, "借用 / " + _returnStatus[_status], ''))
            _historyRecords.insert('', 0, values=(_studentID_value, _name, _department, "借用 / " + _returnStatus[_status], ''))
            _allNameList.save("所有系所名單.xlsx")
        else:
            show_customWindow("該學生未借用！", "error")
    elif int(_borrowCell) == 1:
        if _status == 0:
            show_customWindow("不可重複借用！", "error")
            return
        current_status = _targetSheet.cell(row=_targetRow, column=6).value
        if current_status is not None and current_status != 0:
            show_customWindow("不能重複歸還！", "error")
            return
        for row in _targetRecord.get_children():
            _targetRecord.delete(row)
        _targetSheet.cell(row=_targetRow, column=6).value = _status
        _targetSheet.cell(row=_targetRow, column=7).value = _time
        _targetRecord.insert('', 'end', values=(_studentID_value, _name, _department, "借用 / " + _returnStatus[_status], _time))
        _historyRecords.insert('', 0, values=(_studentID_value, _name, _department, "借用 / " + _returnStatus[_status], _time))
        _allNameList.save("所有系所名單.xlsx")
    else:
        show_customWindow("該學生不在借用狀態！", "error")
    return

def clear_status(_targetRecord, _historyRecords, _status):
    if _targetRow == -1:
        show_customWindow("請輸入學號！", "error")
        return

    _borrowCell = _targetSheet.cell(row=_targetRow, column=5).value
    _statusCell = _targetSheet.cell(row=_targetRow, column=6).value
    _studentID_value = str(_targetSheet.cell(row=_targetRow, column=1).value).strip()
    _name = _targetSheet.cell(row=_targetRow, column=2).value
    _department = get_deparmentInfo(_targetSheet.cell(row=_targetRow, column=3).value, "name")

    if _borrowCell is None:
        show_customWindow("不可刪除未借用！", "error")
        return

    # 刪除「借用」紀錄
    if _status == 0:
        if _statusCell is not None and _statusCell != 0:
            show_customWindow("已歸還，不可刪除借用！", "error")
            return

        # 合法刪除借用
        for row in _targetRecord.get_children():
            _targetRecord.delete(row)

        _targetSheet.cell(row=_targetRow, column=5).value = None
        _targetSheet.cell(row=_targetRow, column=6).value = None
        _targetSheet.cell(row=_targetRow, column=7).value = None

        msg = "已刪除借用"
        _targetRecord.insert('', 'end', values=(_studentID_value, _name, _department, msg, ''))
        _historyRecords.insert('', 0, values=(_studentID_value, _name, _department, msg, ''))

    # 刪除「歸還方式」紀錄
    else:
        if _statusCell is None or _statusCell == 0:
            show_customWindow("未歸還，不可刪除歸還狀態！", "error")
            return

        if _statusCell != _status:
            show_customWindow("錯誤刪除歸還方式！", "error")
            return

        # 合法刪除該歸還方式
        for row in _targetRecord.get_children():
            _targetRecord.delete(row)

        _targetSheet.cell(row=_targetRow, column=6).value = 0
        _targetSheet.cell(row=_targetRow, column=7).value = None

        msg = "借用 / 已刪除" + _deleteStatus[_status - 1]
        _targetRecord.insert('', 'end', values=(_studentID_value, _name, _department, msg, ''))
        _historyRecords.insert('', 0, values=(_studentID_value, _name, _department, msg, ''))

    _allNameList.save("所有系所名單.xlsx")

def get_deparmentInfo(formula, mode):
    if not formula.startswith('=RIGHT('):
        return None
    try:
        content = formula[7:-1]
        text_part, number_part = content.split(',')
        text = text_part.strip().strip('"')
        n = int(number_part.strip())
        if mode == "name":
            return text[-n:]
        elif mode == "depIndex":
            return text[0:4]
        elif mode == "digit":
            return int(text_part[1])
    except Exception as e:
        print(f"解析錯誤: {e}")
        return None

def show_customWindow(message, type):
    window = tk.Toplevel()
    window_width = int(500 * 0.9 / scaling_factor)
    window_height = int(300 * 0.9 / scaling_factor)
    center_x = int(_screenWidth / 2 - window_width / 2)
    center_y = int(_screenHeight / 2 - window_height / 2)
    window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    window.configure(bg="white")
    if type == "error":
        window.title("錯誤！")
        _tempColor = "red"
    elif type == "notice":
        window.title("通知！")
        _tempColor = "blue"
    label = tk.Label(window, text=message, fg=_tempColor, font=font_style, bg="white")
    label.pack(pady=int(50 * 0.9 / scaling_factor))
    ok_button = tk.Button(window, text="確定", font=font_style, command=window.destroy)
    ok_button.pack(pady=int(50 * 0.9 / scaling_factor))
    window.focus_force()
    window.bind('<Return>', lambda event: window.destroy())
    window.grab_set()

def output_borrow_file():
    _collegeFile = openpyxl.Workbook()
    _departmentFile = openpyxl.Workbook()
    _collegeSheet = _collegeFile.active
    _departmentSheet = _departmentFile.active
    _collegeSheet.title = "各院借用統計"
    _departmentSheet.title = "各系借用統計"
    _departmentDigit= []
    _departmentName = []
    _totalNumbersDpt = []
    _temp_total = 0
    order = [0, 2, 4, 3, 1, 5, 6, 8]
    resetNums()
    for irow in range(1, _targetSheet.max_row + 1):
        _tempDepartment = get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "depIndex")
        _firstDigit = _targetSheet.cell(row=irow, column=1).value[0]
        _collegeDigit = get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "digit")
        _studentID = _targetSheet.cell(row=irow, column=1).value
        _tempDegree = -1
        if _tempDepartment and (_firstDigit, _tempDepartment) not in _departmentDigit:
            _departmentDigit.append((_firstDigit, _tempDepartment))
            _departmentName.append(get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "name"))
            _totalNumbersDpt.append(_temp_total)
            _temp_total = 0
        if _targetSheet.cell(row=irow, column=5).value == 1:
            if _studentID and str(_studentID)[0] == '4':
                _tempDegree = 0  # 大學部
            elif _studentID and str(_studentID)[0] in ['5', '6']:
                _tempDegree = 1  # 碩專班
            elif _studentID and str(_studentID)[0] == '8':
                _tempDegree = 2  # 博班
            if _tempDegree >= 0 and _collegeDigit is not None and 1 <= _collegeDigit <= len(_collegeNums):
                _collegeNums[_collegeDigit - 1][_tempDegree] += 1
                _totalNums[_tempDegree] += 1
                _temp_total += 1
    _totalNumbersDpt.append(_temp_total)
    
    for i in range(ord('A'), ord('E') + 1):
        _collegeSheet.column_dimensions[chr(i)].width = 18.5
    for i in range(1, 13):
        _collegeSheet.row_dimensions[i].height = 25
    _collegeSheet.append(["各院人數借用統計"])
    _collegeSheet.append(['', '大學部', '碩專班', '博班'])
    for i in order:
        if _collegeName[i] is not None:
            _collegeSheet.append([_collegeName[i], _collegeNums[i][0], _collegeNums[i][1], _collegeNums[i][2]])
    _collegeSheet.append(["總和", _totalNums[0], _totalNums[1], _totalNums[2]])
    _collegeSheet.append(["借用總和", _totalNums[0] + _totalNums[1] + _totalNums[2]])
    for row in _collegeSheet.iter_rows(min_col=1, max_col=_collegeSheet.max_column, min_row=1):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
    _collegeSheet.merge_cells('A1:D1')
    _collegeSheet.merge_cells('B12:D12')

    _departmentSheet.column_dimensions['A'].width = 55
    _departmentSheet.row_dimensions[1].height = 25
    _departmentSheet.row_dimensions[2].height = 25
    for i in range(3, len(_departmentName) + 3):
        _departmentSheet.row_dimensions[i].height = 18
    _departmentSheet.append(["各系人數借用統計"])
    _departmentSheet.append(["系所名稱", "人數"])
    _departmentSheet.merge_cells('A1:B1')
    for i in range(0, len(_departmentName)):
        _departmentSheet.append([_departmentName[i], _totalNumbersDpt[i+1]])
    for row in _departmentSheet.iter_rows(min_col=1, max_col=_departmentSheet.max_column, min_row=1):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
    create_path("借用統計人數")
    _collegeFile.save("借用統計人數/借用統計人數(學院).xlsx")
    _departmentFile.save('借用統計人數/借用統計人數(系).xlsx')
    show_customWindow("輸出成功！", "notice")

def show_totalStatus():
    today = datetime.today()
    default_month = str(today.month)
    default_day = str(today.day)
    month = _monthVar.get().strip() or default_month
    day = _dayVar.get().strip() or default_day
    period = _selectedOption.get()

    try:
        month = int(month)
        day = int(day)
        if not (1 <= month <= 12) or not (1 <= day <= 31):
            show_customWindow("月份或日期無效！", "error")
            return
    except ValueError:
        show_customWindow("請輸入有效的月份和日期（數字）！", "error")
        return

    degrees = ['學士', '碩士', '博士']
    statuses = ['歸還', '遺失', '新衣', '報廢']
    colleges = ['文', '社', '管', '工', '理', '法', '教', '不分']
    stats = {
        degree: {
            status: {college: 0 for college in colleges}
            for status in statuses
        }
        for degree in degrees
    }

    for row in range(1, _targetSheet.max_row + 1):
        if _targetSheet.cell(row=row, column=5).value != 1:
            continue
        status_value = _targetSheet.cell(row=row, column=6).value
        if status_value is None or status_value == 0:
            continue
        if status_value not in [1, 2, 3, 4]:
            continue
        status = statuses[status_value - 1]
        time_value = _targetSheet.cell(row=row, column=7).value
        if not time_value:
            continue
        if isinstance(time_value, str):
            try:
                time_value = datetime.strptime(time_value, "%Y-%m-%d %H:%M:%S.%f")
            except ValueError:
                try:
                    time_value = datetime.strptime(time_value, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
        if time_value.month != month or time_value.day != day:
            continue
        hour = time_value.hour
        if period == "上午" and hour >= 13:
            continue
        if period == "下午" and hour < 13:
            continue
        student_id = str(_targetSheet.cell(row=row, column=1).value).strip()
        if student_id[0] == '4':
            degree = '學士'
        elif student_id[0] in ['5', '6']:
            degree = '碩士'
        elif student_id[0] == '8':
            degree = '博士'
        else:
            continue
        college_digit = get_deparmentInfo(_targetSheet.cell(row=row, column=3).value, "digit")
        if college_digit is None or not (1 <= college_digit <= 9):
            continue
        college = colleges[college_digit - 1] if college_digit != 8 else colleges[7]
        stats[degree][status][college] += 1

    _window = tk.Toplevel()
    _window.title(f"歸還統計 - {month}月{day}日 {period}")
    window_width = int(_screenWidth * 0.9)  # Fixed 80% of screen width
    window_height = int(_screenHeight * 0.8)  # Fixed 80% of screen height
    position_right = int((_screenWidth - window_width) / 2)  # Center horizontally
    position_down = int((_screenHeight - window_height) / 2)  # Center vertically
    _window.geometry(f"{window_width}x{window_height}+{position_right}+{position_down}")

    canvas = tk.Canvas(_window, bg="white")
    v_scrollbar = tk.Scrollbar(_window, orient="vertical", command=canvas.yview)
    h_scrollbar = tk.Scrollbar(_window, orient="horizontal", command=canvas.xview)
    canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
    v_scrollbar.pack(side="right", fill="y")
    h_scrollbar.pack(side="bottom", fill="x")
    canvas.pack(expand=True, fill="both")

    cell_width = int(window_width / 10)  # 10 columns
    cell_height = int(window_height / 13.5)  # Adjusted for 12 rows to fit content
    padding = 0  # Fixed padding for consistency
    canvas_height = cell_height * (1 + 4 * len(degrees))  # Total rows needed
    canvas.configure(scrollregion=(0, 0, window_width, canvas_height))

    colleges_display = ['學位', '狀態', '文', '社', '管', '工', '理', '法', '教', '不分']
    ordered_indices = [0, 2, 4, 3, 1, 5, 6, 7]
    ordered_colleges = [colleges[i] for i in ordered_indices]

    base_font_size_canvas = 16  # Base font size for canvas
    font_size_header = int(base_font_size_canvas / scaling_factor * 1.2)  # Adjusted for scaling
    font_size_degree = int(base_font_size_canvas / scaling_factor * 1.1)
    font_size_cell = int(base_font_size_canvas / scaling_factor)

    for i, college in enumerate(colleges_display):
        x1 = i * cell_width + padding
        x2 = (i + 1) * cell_width + padding
        y1 = 0 + padding
        y2 = cell_height + padding
        canvas.create_rectangle(x1, y1, x2, y2, outline="black")
        canvas.create_text((x1 + x2) / 2, (y1 + y2) / 2, text=college, font=("微軟正黑體", font_size_header))

    current_row = 1
    for degree in degrees:
        x1 = 0 + padding
        x2 = 10 * cell_width + padding
        y1 = current_row * cell_height + padding
        y2 = (current_row + 4) * cell_height + padding
        canvas.create_rectangle(x1, y1, x2, y2, outline="black", width=3)
        x1 = 0 + padding
        x2 = cell_width + padding
        center_y = (y1 + y2) / 2
        canvas.create_rectangle(x1, y1, x2, y2, outline="black")
        canvas.create_text((x1 + x2) / 2, center_y, text=degree, font=("微軟正黑體", font_size_degree))

        for status in statuses:
            x1 = cell_width + padding
            x2 = 2 * cell_width + padding
            y1 = current_row * cell_height + padding
            y2 = (current_row + 1) * cell_height + padding
            bg_color = "#e6e6e6" if current_row % 2 == 1 else "white"
            canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill=bg_color)
            canvas.create_text((x1 + x2) / 2, (y1 + y2) / 2, text=status, font=("微軟正黑體", font_size_cell))

            for col_idx, college in enumerate(ordered_colleges):
                x1 = (col_idx + 2) * cell_width + padding
                x2 = (col_idx + 3) * cell_width + padding
                canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill=bg_color)
                count = stats[degree][status][college]
                canvas.create_text((x1 + x2) / 2, (y1 + y2) / 2, text=str(count), font=("微軟正黑體", font_size_cell))
            current_row += 1

def integrate_returnNamelist():
    folder_path = "整合名單"
    if not os.path.isdir(folder_path):
        show_customWindow("整合名單資料夾不存在！", "error")
        return

    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and f != "所有系所名單.xlsx"]
    if not excel_files:
        show_customWindow("資料夾中沒有可用的 Excel 檔案！", "error")
        return

    # 建立 student_id → 最新記錄 的 dict
    integrated_records = {}

    for file_name in excel_files:
        try:
            wb = load_workbook(os.path.join(folder_path, file_name))
            sheet = wb.active
            for row in range(2, sheet.max_row + 1):
                student_id = sheet.cell(row=row, column=1).value
                if student_id is None:
                    continue
                student_id = str(student_id).strip()
                borrow_status = sheet.cell(row=row, column=5).value
                return_status = sheet.cell(row=row, column=6).value
                return_time = sheet.cell(row=row, column=7).value

                if borrow_status != 1:
                    continue

                # 解析時間
                parsed_time = None
                if return_time:
                    if isinstance(return_time, str):
                        try:
                            parsed_time = datetime.strptime(return_time, "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            continue
                    elif isinstance(return_time, datetime):
                        parsed_time = return_time

                # 比對是否已有該學號的記錄，若無或時間較新才更新
                if student_id not in integrated_records:
                    integrated_records[student_id] = {
                        'status': return_status,
                        'time': parsed_time
                    }
                else:
                    existing_time = integrated_records[student_id]['time']
                    if parsed_time and (existing_time is None or parsed_time > existing_time):
                        integrated_records[student_id] = {
                            'status': return_status,
                            'time': parsed_time
                        }

        except Exception as e:
            show_customWindow(f"處理檔案 {file_name} 時發生錯誤：{str(e)}", "error")
            continue

    # 🔍 對 _allNameList 的每一筆進行更新
    updated = False
    for row in range(2, _targetSheet.max_row + 1):
        student_id = _targetSheet.cell(row=row, column=1).value
        if student_id is None:
            continue
        student_id = str(student_id).strip()

        if student_id in integrated_records:
            record = integrated_records[student_id]
            _targetSheet.cell(row=row, column=5).value = 1
            _targetSheet.cell(row=row, column=6).value = record['status']
            _targetSheet.cell(row=row, column=7).value = record['time'].strftime("%Y-%m-%d %H:%M:%S") if record['time'] else None
            updated = True

    if updated:
        _allNameList.save("所有系所名單.xlsx")

        # ✅ 改為移動檔案到備份資料夾（避免防毒誤報）
        backup_path = os.path.join(folder_path, "已整合備份")
        os.makedirs(backup_path, exist_ok=True)

        for file_name in excel_files:
            try:
                source = os.path.join(folder_path, file_name)
                destination = os.path.join(backup_path, file_name)
                shutil.move(source, destination)
            except Exception as e:
                show_customWindow(f"移動檔案 {file_name} 時發生錯誤：{str(e)}", "error")

        show_customWindow("名單整合成功，整合檔案已移至『已整合備份』資料夾！", "notice")
    else:
        show_customWindow("沒有需要更新的記錄！", "notice")

def outputfile():
    today = datetime.today()
    default_month = str(today.month)
    default_day = str(today.day)
    month = _monthVar.get().strip() or default_month
    day = _dayVar.get().strip() or default_day
    period = _selectedOption.get()

    if not month or not day:
        show_customWindow("請輸入月份和日期！", "error")
        return
    try:
        month = int(month)
        day = int(day)
        if not (1 <= month <= 12) or not (1 <= day <= 31):
            show_customWindow("月份或日期無效！", "error")
            return
    except ValueError:
        show_customWindow("請輸入有效的月份和日期（數字）！", "error")
        return

    target_date = f"{today.year}-{month:02d}-{day:02d} 00:00:00"
    day = datetime.strptime(target_date, '%Y-%m-%d %H:%M:%S')
    resetNums()
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    order = [0, 2, 4, 3, 1, 5, 6, 8]  # 不含 index 7 (None)
    time_1 = timedelta(0)
    time_2 = timedelta(hours=13)
    time_3 = timedelta(days=1)

    degrees = ['學士', '碩士', '博士']
    colleges = ['文', '社', '管', '工', '理', '法', '教', None, '不分']
    non_returned_stats = {
        degree: {status: {college: 0 for college in colleges if college is not None} for status in ['遺失', '新衣', '報廢']}
        for degree in degrees
    }
    total_counts = {
        degree: {college: 0 for college in colleges if college is not None}
        for degree in degrees
    }

    for i in range(ord('A'), ord('D')+1):
        sheet2.column_dimensions[chr(i)].width = 23
    for i in range(1, 11):
        sheet2.row_dimensions[i].height = 24
    for i in range(12, 22):
        sheet2.row_dimensions[i].height = 24
    for i in range(23, 33):
        sheet2.row_dimensions[i].height = 24

    student_data = {i: [] for i in order}
    record_count = 0

    for row in range(1, _targetSheet.max_row + 1):
        code = str(_targetSheet.cell(row=row, column=3).value).lstrip('=RIGHT("')
        return_time = _targetSheet.cell(row=row, column=7).value
        if return_time is None:
            continue
        try:
            if isinstance(return_time, str):
                try:
                    return_time = datetime.strptime(return_time, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        return_time = datetime.strptime(return_time, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        continue
            time_diff = return_time - day
        except Exception:
            continue

        if ((time_diff > time_1 and time_diff < time_2 and period == "上午") or 
            (time_diff > time_2 and time_diff < time_3 and period == "下午") or 
            (time_diff > time_1 and time_diff < time_3 and period == "全天")):
            try:
                college_idx = int(code[0]) - 1 if code and code[0].isdigit() else -1
                if college_idx not in order or college_idx >= len(colleges):
                    continue
                college = colleges[college_idx]
                if college is None:
                    continue
                idx = order.index(college_idx)
                student_id = _targetSheet.cell(row=row, column=1).value or "未知"
                name = _targetSheet.cell(row=row, column=2).value or "無姓名"
                department = get_deparmentInfo(_targetSheet.cell(row=row, column=3).value, "name") or "未知系所"
                phone = _targetSheet.cell(row=row, column=4).value or "無電話"
                status_value = _targetSheet.cell(row=row, column=6).value
                student_data[order[idx]].append([student_id, name, department, phone])
                record_count += 1

                if str(student_id)[0] == '4':
                    degree = '學士'
                    _collegeNums[order[idx]][0] += 1
                    _totalNums[0] += 1
                elif str(student_id)[0] in ['5', '6']:
                    degree = '碩士'
                    _collegeNums[order[idx]][1] += 1
                    _totalNums[1] += 1
                elif str(student_id)[0] == '8':
                    degree = '博士'
                    _collegeNums[order[idx]][2] += 1
                    _totalNums[2] += 1
                else:
                    continue  # 無法辨識學位

                total_counts[degree][college] += 1
                if status_value in [2, 3, 4]:
                    status = ['遺失', '新衣', '報廢'][status_value - 2]
                    non_returned_stats[degree][status][college] += 1
            except (ValueError, IndexError, KeyError):
                continue

    # 原始統計
    # 第一表：所有狀態（含正常與非正常）統計
    sheet2.append([f'{day.strftime("%m/%d")}' ,'學士', '碩士', '博士'])
    for idx in order:
        college = colleges[idx]
        if college is None:
            continue
        sheet2.append([
            _collegeName[idx],
            total_counts['學士'][college],
            total_counts['碩士'][college],
            total_counts['博士'][college],
        ])
    sheet2.append([
        '總和',
        sum(total_counts['學士'].values()),
        sum(total_counts['碩士'].values()),
        sum(total_counts['博士'].values()),
    ])
    sheet2.append([])


    # 非歸還狀態統計
    # 第二表：非歸還狀態統計（遺失、新衣、報廢）
    sheet2.append(['','學士(遺/新/廢)', '碩士(遺/新/廢)', '博士(遺/新/廢)'])
    for idx in order:
        college = colleges[idx]
        if college is None:
            continue
        row_data = [_collegeName[idx]]
        for degree in degrees:
            counts = [non_returned_stats[degree][s][college] for s in ['遺失', '新衣', '報廢']]
            row_data.append(f"{counts[0]} / {counts[1]} / {counts[2]}")
        sheet2.append(row_data)

    # 小計
    sheet2.append(['小計'] + [
        f"{sum(non_returned_stats[degree]['遺失'].values())} / "
        f"{sum(non_returned_stats[degree]['新衣'].values())} / "
        f"{sum(non_returned_stats[degree]['報廢'].values())}"
        for degree in degrees
    ])
    sheet2.append([])


    # 最終有效歸還統計（總數扣除非歸還）
    # 第三表：實際有效歸還（總數 - 非歸還）
    sheet2.append(['學院', '學士', '碩士', '博士'])
    final_total = {'學士': 0, '碩士': 0, '博士': 0}
    for idx in order:
        college = colleges[idx]
        if college is None:
            continue
        row = [_collegeName[idx]]
        for degree in degrees:
            total = total_counts[degree][college]
            lost = non_returned_stats[degree]['遺失'][college]
            new = non_returned_stats[degree]['新衣'][college]
            scrap = non_returned_stats[degree]['報廢'][college]
            valid = total - lost - new - scrap
            final_total[degree] += valid
            row.append(valid)
        sheet2.append(row)

    sheet2.append([
        '總和',
        final_total['學士'],
        final_total['碩士'],
        final_total['博士']
    ])
    # 學生名單區
    blank_rows = []
    for idx in order:
        if student_data[idx]:
            sheet2.append([""])
            blank_rows.append(sheet2.max_row)
            sheet2.append([_collegeName[idx]])
            current_row = sheet2.max_row
            sheet2.merge_cells(f'A{current_row}:D{current_row}')
            sheet2[f'A{current_row}'].alignment = center_alignment
            sheet2[f'A{current_row}'].border = thin_border
            sheet2.append(["學號", "姓名", "系所", "電話"])
            for student in student_data[idx]:
                sheet2.append(student)

    for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, max_col=4):
        if row[0].row in blank_rows:
            continue
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.border = thin_border

    if record_count == 0:
        show_customWindow("指定日期和時段無歸還記錄！", "error")
        wb2.close()
        return

    create_path("當日歸還統計資料")
    if period == "上午":
        filename = f"當日歸還統計資料/{day.strftime('%m.%d')}上午歸還統計.xlsx"
    elif period == "下午":
        filename = f"當日歸還統計資料/{day.strftime('%m.%d')}下午歸還統計.xlsx"
    else:
        filename = f"當日歸還統計資料/{day.strftime('%m.%d')}全天歸還統計.xlsx"

    wb2.save(filename)
    show_customWindow("輸出成功！", "notice")

def output_noReturn_file():
    resetNums()
    wb2 = openpyxl.Workbook()
    wb3 = openpyxl.Workbook()
    wb4 = openpyxl.Workbook()
    wb5 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet3 = wb3.active
    sheet4 = wb4.active
    sheet5 = wb5.active
    department = []
    _departmentDigit = []
    number_people = []
    _temp_total = 0
    day = datetime.now()
    order = [0, 2, 4, 3, 1, 5, 6, 8]
    sheet2.column_dimensions['A'].width = 10.5
    sheet2.column_dimensions['B'].width = 10
    sheet2.column_dimensions['C'].width = 50
    sheet2.column_dimensions['D'].width = 11
    sheet3.column_dimensions['A'].width = 8
    sheet3.column_dimensions['B'].width = 12
    sheet3.column_dimensions['C'].width = 12
    sheet3.column_dimensions['D'].width = 8
    sheet2.append(['學號', '姓名', '系所名稱', '電話'])
    for irow in range(1, _targetSheet.max_row + 1):
        student_id = _targetSheet.cell(row=irow, column=1).value
        _tempDepartment = get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "depIndex")
        _firstDigit = _targetSheet.cell(row=irow, column=1).value[0]
        name = _targetSheet.cell(row=irow, column=2).value
        department_name = get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "name")
        phone = _targetSheet.cell(row=irow, column=4).value
        college_index = get_deparmentInfo(_targetSheet.cell(row=irow, column=3).value, "digit")
        if _tempDepartment and (_firstDigit, _tempDepartment) not in _departmentDigit:
            _departmentDigit.append((_firstDigit, _tempDepartment))
            department.append(department_name)
            number_people.append(_temp_total)
            _temp_total = 0
        if _targetSheet.cell(row=irow, column=5).value == 1 and _targetSheet.cell(row=irow, column=6).value is None:
            sheet2.append([student_id, name, department_name, phone])
            sheet3.append(['10', student_id, '學位服未歸還', 'N'])
            if str(student_id)[0] == '4':
                _collegeNums[college_index - 1][0] += 1
                _totalNums[0] += 1
                _temp_total += 1
            elif str(student_id)[0] in ['5', '6']:
                _collegeNums[college_index-1][1] += 1
                _totalNums[1] += 1
                _temp_total += 1
            elif str(student_id)[0] == '8':
                _collegeNums[college_index-1][2] += 1
                _totalNums[2] += 1
                _temp_total += 1
    number_people.append(_temp_total)
    for i in range(ord('A'), ord('D')+1):
        sheet4.column_dimensions[chr(i)].width = 23
    for i in range(1, 13):
        sheet4.row_dimensions[i].height = 25
    sheet4.append(["未歸還統計人數"])
    sheet4.append([str(day.strftime('%m/%d')), '學士', '碩士', '博士'])
    for i in order:
        sheet4.append([_collegeName[i], _collegeNums[i][0], _collegeNums[i][1], _collegeNums[i][2]])
    sheet4.append(["總和", _totalNums[0], _totalNums[1], _totalNums[2]])
    sheet4.append(["未歸還總和", _totalNums[0] + _totalNums[1] + _totalNums[2]])
    sheet4.merge_cells('A1:D1')
    sheet4.merge_cells('B12:D12')
    for row in sheet2.iter_rows(min_row=1, max_col=4):
        for cell in row:
            sheet2[cell.coordinate].alignment = center_alignment
    for row in sheet3.iter_rows(min_row=1, max_col=4):
        for cell in row:        
            sheet3[cell.coordinate].alignment = center_alignment
    for row in sheet4.iter_rows(min_row=1, max_col=4):
        for cell in row:
            sheet4[cell.coordinate].alignment = center_alignment
            sheet4[cell.coordinate].border = thin_border
    sheet5.column_dimensions['A'].width = 55
    sheet5.row_dimensions[1].height = 25
    sheet5.row_dimensions[2].height = 25
    for i in range(3, len(department) + 3):
        sheet5.row_dimensions[i].height = 18
    sheet5.append([str(day.strftime('%m/%d')) + '未歸還統計'])
    sheet5.append(["系所名稱", "人數"])
    sheet5.merge_cells('A1:B1')
    for i in range(len(department)):
        sheet5.append([department[i], number_people[i+1]])
    for row in sheet5.iter_rows(min_row=1, max_col=2):
        for cell in row:
            sheet5[cell.coordinate].alignment = center_alignment
            sheet5[cell.coordinate].border = thin_border
    create_path("未歸還名單")
    wb2.save(f"未歸還名單/{day.strftime('%m.%d')}未歸還名單.xlsx")
    wb3.save(f"未歸還名單/{day.strftime('%m.%d')}離校系統未歸還名冊.xlsx")
    wb4.save(f"未歸還名單/{day.strftime('%m.%d')}未歸還統計人數(學院).xlsx")
    wb5.save(f"未歸還名單/{day.strftime('%m.%d')}未歸還統計人數(系).xlsx")
    show_customWindow("輸出成功！", "notice")

def sort_history_record(_historyRecords, key, reverse):
    valid_keys = ["學號", "姓名", "系所", "狀態", "時間"]
    if key not in valid_keys:
        show_customWindow("無效的排序欄位！", "error")
        return
    data_list = [(_historyRecords.set(k, key), k) for k in _historyRecords.get_children('')]
    if key == "時間":
        data_list.sort(key=lambda t: (t[0] == '', t[0]), reverse=reverse)
    else:
        data_list.sort(key=lambda t: t[0], reverse=reverse)
    for index, (val, k) in enumerate(data_list):
        _historyRecords.move(k, '', index)
    _historyRecords.heading(key, command=lambda: sort_history_record(_historyRecords, key, not reverse))

def main():
    init_excel()
    frame_width = _screenWidth * 0.85
    base_width = int(frame_width / 26 * 0.9 / scaling_factor)

    # 學號輸入區
    top_frame = tk.Frame(_mainWindow, bg=_backgroundColor)
    top_frame.grid(row=0, column=0, sticky='nsew')
    top_frame.columnconfigure(0, weight=0)
    top_frame.columnconfigure(1, weight=1)
    top_frame.columnconfigure(2, weight=0)
    top_frame.columnconfigure(3, weight=1)
    top_frame.columnconfigure(4, weight=0)
    top_frame.columnconfigure(5, weight=0)
    top_frame.columnconfigure(6, weight=0)
    top_frame.columnconfigure(7, weight=0)
    top_frame.columnconfigure(8, weight=0)
    top_frame.columnconfigure(9, weight=0)
    top_frame.columnconfigure(10, weight=0)
    top_frame.rowconfigure(0, weight=0)
    top_frame.rowconfigure(1, weight=1)
    top_frame.rowconfigure(2, weight=0)
    top_frame.grid_propagate(False)
    top_frame.configure(height=int(_screenHeight * 0.23))

    _studentID_label = tk.Label(top_frame, bg=_backgroundColor, text="學號：", fg='white', font=header_font, width=5, height=2)
    _studentID_label.grid(column=0, row=0, sticky='w', padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    studentID_input = tk.Entry(top_frame, textvariable=ID, bd=3, font=font_style)
    studentID_input.grid(column=1, row=0, sticky='ew', pady=int(20 * 0.9 / scaling_factor))
    studentID_input.bind('<Return>', lambda event: search_studentID(_targetRecord))
    studentID_input.focus_set()
    _searchButton = tk.Button(top_frame, text="查詢", font=header_font, command=lambda: search_studentID(_targetRecord), bg=_buttonColor, fg='white', width=10)
    _searchButton.grid(column=2, row=0, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _displayTargetInfo = ("學號", "姓名", "系所", "狀態", "時間")
    _targetRecord = ttk.Treeview(top_frame, height=1, show="headings", columns=_displayTargetInfo)
    _targetRecord.grid(column=0, row=1, columnspan=11, sticky='nsew', padx=int(20 * 0.9 / scaling_factor), pady=int(15 * 0.9 / scaling_factor))
    total_width = int(frame_width * 0.95)
    column_ratios = [3, 4, 5, 3, 4]
    total_ratio = sum(column_ratios)
    for i, col in enumerate(_displayTargetInfo):
        _targetRecord.column(col, width=int(total_width * column_ratios[i] / total_ratio), anchor='center')
        _targetRecord.heading(col, text=col)
    _statusButton1 = tk.Radiobutton(top_frame, text="借用", variable=_statusOption1, value=0, font=header_font, bg=_backgroundColor, fg='white', selectcolor=_backgroundColor)
    _statusButton2 = tk.Radiobutton(top_frame, text="歸還", variable=_statusOption1, value=1, font=header_font, bg=_backgroundColor, fg='white', selectcolor=_backgroundColor)
    _statusButton3 = tk.Radiobutton(top_frame, text="遺失", variable=_statusOption1, value=2, font=header_font, bg=_backgroundColor, fg='white', selectcolor=_backgroundColor)
    _statusButton4 = tk.Radiobutton(top_frame, text="新衣", variable=_statusOption1, value=3, font=header_font, bg=_backgroundColor, fg='white', selectcolor=_backgroundColor)
    _statusButton5 = tk.Radiobutton(top_frame, text="報廢", variable=_statusOption1, value=4, font=header_font, bg=_backgroundColor, fg='white', selectcolor=_backgroundColor)
    _statusButton1.grid(column=4, row=0, padx=int(15 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor), sticky='e')
    _statusButton2.grid(column=5, row=0, padx=int(15 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor), sticky='e')
    _statusButton3.grid(column=6, row=0, padx=int(15 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor), sticky='e')
    _statusButton4.grid(column=7, row=0, padx=int(15 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor), sticky='w')
    _statusButton5.grid(column=8, row=0, padx=int(15 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor), sticky='w')
    _confirmButton = tk.Button(top_frame, text="確認", command=lambda: change_return_status(_targetRecord, _historyRecords, _statusOption1.get()), font=header_font, bg=_buttonColor, fg='white', width=8)
    _confirmButton.grid(column=9, row=0, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _clearStatusButton = tk.Button(top_frame, text="刪除", font=header_font, command=lambda: clear_status(_targetRecord, _historyRecords, _statusOption1.get()), bg=_buttonColor, fg='white', width=8)
    _clearStatusButton.grid(column=10, row=0, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    separator1 = tk.Frame(top_frame, bg="white", height=3)
    separator1.grid(column=0, row=3, columnspan=11, sticky='nsew')
    separator1.grid_propagate(False)

    # 歷史記錄區
    middle_frame = tk.Frame(_mainWindow, bg=_backgroundColor)
    middle_frame.grid(row=1, column=0, sticky='nsew')
    middle_frame.columnconfigure(0, weight=0)
    middle_frame.columnconfigure(1, weight=0)
    middle_frame.columnconfigure(2, weight=0)
    middle_frame.columnconfigure(3, weight=0)
    middle_frame.columnconfigure(4, weight=0)
    middle_frame.columnconfigure(5, weight=0)
    middle_frame.columnconfigure(6, weight=1)
    middle_frame.columnconfigure(7, weight=0)
    middle_frame.rowconfigure(0, weight=0)
    middle_frame.rowconfigure(1, weight=0)
    middle_frame.rowconfigure(2, weight=1)
    middle_frame.grid_propagate(False)
    middle_frame.configure(height=int(_screenHeight * 0.64))

    _record_label = tk.Label(middle_frame, bg=_backgroundColor, text="歷史記錄：", fg='white', font=header_font, width=8)
    _record_label.grid(column=0, row=0, columnspan=2, sticky='w', padx=int(20 * 0.9 / scaling_factor), pady=int(10 * 0.9 / scaling_factor))
    _monthLabel = tk.Label(middle_frame, bg=_backgroundColor, text="月：", fg='white', font=header_font, width=3)
    _monthLabel.grid(column=0, row=1, sticky='w', padx=(int(20 * 0.9 / scaling_factor), int(10 * 0.9 / scaling_factor)))
    _monthInput = tk.Entry(middle_frame, textvariable=_monthVar, bd=3, font=font_style, width=6)
    _monthInput.grid(column=1, row=1, sticky='w')
    _dayLabel = tk.Label(middle_frame, bg=_backgroundColor, text="日：", fg='white', font=header_font, width=3)
    _dayLabel.grid(column=2, row=1, sticky='w', padx=int(20 * 0.9 / scaling_factor))
    _dayInput = tk.Entry(middle_frame, textvariable=_dayVar, bd=3, font=font_style, width=6)
    _dayInput.grid(column=3, row=1, sticky='w')
    _dropdownMenu = tk.OptionMenu(middle_frame, _selectedOption, *_statusOptions)
    _dropdownMenu.config(font=header_font, width=6, bg='white', fg='black', relief='groove')
    _dropdownMenu['menu'].config(font=header_font, activebackground='lightgray', activeforeground='red', fg='black')
    _dropdownMenu.grid(column=4, row=1, sticky='w', padx=int(20 * 0.9 / scaling_factor))
    _buttonforReturnStatistics = tk.Button(middle_frame, text="歸還統計", fg="white", command=show_totalStatus, bg=_buttonColor, font=header_font, width=10)
    _buttonforReturnStatistics.grid(column=5, row=1, padx=int(15 * 0.9 / scaling_factor), pady=int(10 * 0.9 / scaling_factor))

    _historyRecords = ttk.Treeview(middle_frame, height=10, show="headings", columns=_displayTargetInfo)
    scrollbar = tk.Scrollbar(middle_frame, orient="vertical", command=_historyRecords.yview)
    _historyRecords.configure(yscrollcommand=scrollbar.set)
    _historyRecords.grid(column=0, row=2, columnspan=7, sticky='nsew', padx=(int(20 * 0.9 / scaling_factor), 0), pady=int(10 * 0.9 / scaling_factor))
    scrollbar.grid(column=7, row=2, sticky='ns', padx=(0, int(20 * 0.9 / scaling_factor)), pady=int(10 * 0.9 / scaling_factor))
    for i, col in enumerate(_displayTargetInfo):
        _historyRecords.column(col, width=int(total_width * column_ratios[i] / total_ratio), anchor='center')
        _historyRecords.heading(col, text=col, command=lambda c=col: sort_history_record(_historyRecords, c, False))
    separator = tk.Frame(middle_frame, bg="white", height=3)
    separator.grid(column=0, row=3, columnspan=8, sticky='nsew')
    separator.grid_propagate(False)

    # 其他功能按鈕區
    bottom_frame = tk.Frame(_mainWindow, bg=_backgroundColor)
    bottom_frame.grid(row=2, column=0, sticky='nsew')
    bottom_frame.rowconfigure(0, weight=0)
    bottom_frame.rowconfigure(1, weight=1)
    bottom_frame.grid_propagate(False)
    bottom_frame.configure(height=int(_screenHeight * 0.13))

    _othersFunction = tk.Label(bottom_frame, text="其他功能：", fg='white', bg=_backgroundColor, font=header_font)
    _othersFunction.grid(column=0, row=1, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _buttonforBorrow = tk.Button(bottom_frame, text="輸出借用統計", fg="white", command=output_borrow_file, bg=_buttonColor, font=header_font, width=14)
    _buttonforBorrow.grid(column=1, row=1, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _buttonforNoReturn = tk.Button(bottom_frame, text="輸出未還統計", fg="white", command=output_noReturn_file, bg=_buttonColor, font=header_font, width=14)
    _buttonforNoReturn.grid(column=2, row=1, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _buttonforReturn = tk.Button(bottom_frame, text="列印歸還統計", fg="white", command=outputfile, bg=_buttonColor, font=header_font, width=14)
    _buttonforReturn.grid(column=3, row=1, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))
    _buttonforCombine = tk.Button(bottom_frame, text="整合歸還名單", command=integrate_returnNamelist, fg="white", bg=_buttonColor, font=header_font, width=14)
    _buttonforCombine.grid(column=4, row=1, padx=int(20 * 0.9 / scaling_factor), pady=int(20 * 0.9 / scaling_factor))

    _mainWindow.mainloop()

if __name__ == '__main__':
    main()
